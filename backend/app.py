from flask import Flask, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from flask_migrate import Migrate
import os
from datetime import datetime
from dotenv import load_dotenv
import json
from io import BytesIO
import requests

# Optional imports - will be imported only if available
try:
    import pandas as pd
    import openpyxl
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("Warning: pandas/openpyxl not available. Excel upload/download will be disabled.")

try:
    import pymssql
    PYMSSQL_AVAILABLE = True
except ImportError:
    PYMSSQL_AVAILABLE = False
    print("Warning: pymssql not available. SQL Server connections will be disabled.")

# Load environment variables
load_dotenv()

app = Flask(__name__)

#####################################################################
#                    SQL SERVER DATA RETRIEVAL                        #
#####################################################################

# SQL Server Configuration for Performance Data
SQL_SERVER_CONFIG = {
    'server': os.getenv('SQL_SERVER_HOST', 'your_sql_server_host'),
    'database': os.getenv('SQL_SERVER_DATABASE', 'your_database_name'),
    'username': os.getenv('SQL_SERVER_USERNAME', 'your_username'),
    'password': os.getenv('SQL_SERVER_PASSWORD', 'your_password'),
    'driver': os.getenv('SQL_SERVER_DRIVER', '{SQL Server}'),  # or '{ODBC Driver 17 for SQL Server}'
    'port': os.getenv('SQL_SERVER_PORT', '1433')
}

# SQL Server Connection String Template
SQL_SERVER_CONNECTION_STRING = (
    f"DRIVER={SQL_SERVER_CONFIG['driver']};"
    f"SERVER={SQL_SERVER_CONFIG['server']};"
    f"DATABASE={SQL_SERVER_CONFIG['database']};"
    f"UID={SQL_SERVER_CONFIG['username']};"
    f"PWD={SQL_SERVER_CONFIG['password']};"
    f"PORT={SQL_SERVER_CONFIG['port']};"
)

# Servicing Team Query Template - Aggregated for Bonus Calculation
SERVICING_TEAM_QUERY = """
SELECT 
    AssetManager,
    COUNT(*) as total_collections,
    SUM(TotalAmount) as total_amount,
    SUM(CASE WHEN CFType IN ('CF', 'SSA') THEN TotalAmount ELSE 0 END) as cash_flow_amount,
    SUM(CASE WHEN CFType = 'CF' THEN TotalAmount ELSE 0 END) as cf_amount,
    SUM(CASE WHEN CFType = 'SSA' THEN TotalAmount ELSE 0 END) as ssa_amount,
    SUM(CASE WHEN CFType = 'Legal' THEN TotalAmount ELSE 0 END) as legal_amount,
    SUM(CASE WHEN CFType = 'Non-CF' THEN TotalAmount ELSE 0 END) as non_cf_amount,
    COUNT(CASE WHEN CFType IN ('CF', 'SSA') THEN 1 END) as cf_collections_count,
    COUNT(CASE WHEN CFType = 'CF' THEN 1 END) as cf_count,
    COUNT(CASE WHEN CFType = 'SSA' THEN 1 END) as ssa_count,
    COUNT(CASE WHEN CFType = 'Legal' THEN 1 END) as legal_count,
    COUNT(CASE WHEN CFType = 'Non-CF' THEN 1 END) as non_cf_count,
    COUNT(CASE WHEN CFType = 'Non-CF' AND FlagColumn = 0 THEN 1 END) as ncf_count
FROM (
    SELECT 
        ReceivedDate, 
        TotalAmount, 
        CashFlowType, 
        CollectionCategory,
        FlagColumn,
        CASE 
            WHEN CollectionCategory IN (
                'Discounted Payoff  - Unsecured', 
                'Discounted Payoff  - Secured', 
                'Installment', 
                'Workout', 
                'Prepayment Full', 
                'Prepayment Partial', 
                'Consensual Sale Agreement'
            ) THEN 'Amicable Debtor Resolution'
            WHEN CollectionCategory IN (
                'Rent', 
                'Pspa', 
                'Sale Deed'
            ) THEN 'Real Estate'
            WHEN CollectionCategory IN (
                'Assignment Of Award - Sale Third Party', 
                'Collateral Sale', 
                'Loan Sale'
            ) THEN 'Third Part Resolution'
        END AS CollectionCategoryGroup,
        CASE 
            WHEN CollectionCategory IN (
                'Assignment Of Award – Sale Third Party', 
                'Cash In Court Third Party - Sale At Auction', 
                'Cash In Court Third Party - Servicing'
            ) THEN 'SSA'
            WHEN CollectionCategory IN (
                'Rent', 
                'Pspa', 
                'Sale Deed', 
                'Workout', 
                'Prepayment Partial', 
                'Prepayment Full', 
                'Loan Sale', 
                'Installment', 
                'Discounted Payoff  - Secured', 
                'Discounted Payoff  - Unsecured', 
                'Collateral Sale'
            ) THEN 'CF'
            WHEN CollectionCategory IN (
                'Cash In Court', 
                'Cash In Court Third Party - Secured', 
                'Cash In Court Third Party - Unsecured'
            ) THEN 'Legal'
            WHEN CollectionCategory IN (
                'Deed In Lieu', 
                'Consensual Sale Agreement'
            ) OR CollectionCategory IS NULL THEN 'Non-CF'
        END AS CFType,
        CollectionID,
        AssetManager
    FROM 
        CollectionDetail
    WHERE 
        Country = 'ESPANA'
        AND ReceivedDate BETWEEN '{start_date}' AND '{end_date}'
) AS subquery
GROUP BY AssetManager
ORDER BY 
    AssetManager;
"""

# Legal Team Query Template
LEGAL_TEAM_QUERY = """
SELECT DISTINCT 
    l.InternalLawyerName, 
    act.LegalStage, 
    prop.PropertyID, 
    prop.LastValuationAmount, 
    act.Portfolio, 
    act.PortfolioID, 
    act.JudicialProcessID,
    act.LegalActID, 
    act.LegalActCode, 
    act.ActDate, 
    act.ActAmount,
    act.CreationUser,
    act.CreationDate,
    CASE 
        WHEN act.LegalActCode = 'Auction Start Date and Official ID' THEN 'Auction'
        WHEN act.LegalActCode = 'Assigment of awarding celebrated' THEN 'Assigment of awarding'
        WHEN act.LegalActCode IN (
            'Cash In Court Third Party - Secured', 
            'Cash In Court Third Party - Unsecured'
        ) THEN 'Cash In Court'
        WHEN act.LegalActCode = 'Awarding Title' THEN 'Testimony'
        WHEN act.LegalActCode = 'Lawsuit Presentation Date' THEN 'Demands'
        WHEN act.LegalActCode = 'OutCome - Judicial Possession of Keys' THEN 'Possession'
    END AS Bucket
FROM 
    legalactactivity AS act
    INNER JOIN legal AS l ON act.JudicialProcessID = l.JudicialProcessID
    LEFT JOIN Property AS prop ON act.PropertyID = prop.PropertyID
WHERE 
    act.countryID = 2
    AND act.ActDate >= '{start_date}' AND act.ActDate < '{end_date}'
    AND act.CreationDate >= '{start_date}'
    AND act.LegalActID IN (8, 23, 61, 71, 98, 134, 214, 258)
ORDER BY 
    l.InternalLawyerName, 
    act.LegalStage;
"""

#####################################################################
#                    LOCAL APPLICATION LOGIC                          #
#####################################################################

# Initialize Flask app and extensions
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'sqlite:///bonus_calc.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Initialize extensions
db = SQLAlchemy(app)
migrate = Migrate(app, db)
CORS(app)

# Quarter date ranges for local calculations
QUARTER_DATES = {
    'Q1': {'start': '01-01', 'end': '03-31'},
    'Q2': {'start': '04-01', 'end': '06-30'},
    'Q3': {'start': '07-01', 'end': '09-30'},
    'Q4': {'start': '10-01', 'end': '12-31'}
}

#####################################################################
#                    LOCAL DATABASE MODELS                            #
#####################################################################

# Database Models
class Team(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    employees = db.relationship('Employee', backref='team', lazy=True)
    incentive_parameters = db.relationship('IncentiveParameter', backref='team', lazy=True)
    team_member_data = db.relationship('TeamMemberData', backref='team', lazy=True)

class Employee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    surname = db.Column(db.String(100), nullable=False)
    employee_code = db.Column(db.String(50), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    position = db.Column(db.String(100))
    salary = db.Column(db.Float)
    target = db.Column(db.Float, default=0.0)  # New field for bonus calculation target
    team_id = db.Column(db.Integer, db.ForeignKey('team.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    performance_records = db.relationship('PerformanceRecord', backref='employee', lazy=True)

class PerformanceRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=False)
    month = db.Column(db.Integer, nullable=False)
    year = db.Column(db.Integer, nullable=False)
    productivity_score = db.Column(db.Float, default=0.0)
    quality_score = db.Column(db.Float, default=0.0)
    attendance_score = db.Column(db.Float, default=0.0)
    overall_score = db.Column(db.Float, default=0.0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class IncentiveParameter(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    team_id = db.Column(db.Integer, db.ForeignKey('team.id'), nullable=False)
    category = db.Column(db.String(50), nullable=False)  # Analyst, Associate, Senior Associate
    base_bonus = db.Column(db.Float, default=0.0)  # Base bonus percentage
    parameter_name = db.Column(db.String(100), nullable=False)
    base_value = db.Column(db.Float, default=0.0)
    multiplier = db.Column(db.Float, default=1.0)
    min_threshold = db.Column(db.Float, default=0.0)
    max_threshold = db.Column(db.Float, default=100.0)
    is_active = db.Column(db.Boolean, default=True)
    quarter = db.Column(db.String(2), nullable=False)  # Q1, Q2, Q3, Q4
    year = db.Column(db.Integer, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class BonusCalculation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=False)
    month = db.Column(db.Integer, nullable=False)
    year = db.Column(db.Integer, nullable=False)
    quarter = db.Column(db.String(2), nullable=False)  # Q1, Q2, Q3, Q4
    base_salary = db.Column(db.Float, nullable=False)
    performance_score = db.Column(db.Float, nullable=False)
    bonus_amount = db.Column(db.Float, nullable=False)
    calculation_date = db.Column(db.DateTime, default=datetime.utcnow)

class TeamMemberData(db.Model):
    """Model to store detailed team member data from Excel uploads"""
    id = db.Column(db.Integer, primary_key=True)
    team_id = db.Column(db.Integer, db.ForeignKey('team.id'), nullable=False)
    quarter = db.Column(db.String(10), nullable=False)  # Q1, Q2, Q3, Q4
    year = db.Column(db.Integer, nullable=False)
    
    # Employee identification
    employee_name = db.Column(db.String(200), nullable=False)
    employee_code = db.Column(db.String(100), nullable=False)
    category = db.Column(db.String(100))
    team_leader = db.Column(db.String(200))
    
    # Legal Team specific fields
    legal_manager = db.Column(db.String(200))
    employee_hash = db.Column(db.String(100))
    quarterly_incentive = db.Column(db.Float)
    lawsuit_presentation_target = db.Column(db.Float)
    auction_target = db.Column(db.Float)
    cdr_target = db.Column(db.Float)
    testimonies_target = db.Column(db.Float)
    possessions_target = db.Column(db.Float)
    cic_target = db.Column(db.Float)
    
    # Servicing Team specific fields
    asset_sales_manager = db.Column(db.String(200))
    employee_number = db.Column(db.String(100))
    quarter_incentive_base = db.Column(db.Float)
    main_portfolio = db.Column(db.String(200))
    cash_flow = db.Column(db.Float)
    cash_flow_target = db.Column(db.Float)
    ncf = db.Column(db.Float)
    ncf_target = db.Column(db.Float)
    
    # Calculated fields
    cash_flow_percentage = db.Column(db.Float)
    ncf_percentage = db.Column(db.Float)
    incentive_cf = db.Column(db.Float)
    total_incentive = db.Column(db.Float)
    q1_incentive = db.Column(db.Float)
    
    # Legal team calculated fields
    lawsuit_presentation = db.Column(db.Float)
    lawsuit_presentation_percentage = db.Column(db.Float)
    lawsuit_weight = db.Column(db.Float)
    auction = db.Column(db.Float)
    auction_percentage = db.Column(db.Float)
    auction_weight = db.Column(db.Float)
    cdr = db.Column(db.Float)
    cdr_percentage = db.Column(db.Float)
    cdr_weight = db.Column(db.Float)
    testimonies = db.Column(db.Float)
    testimonies_percentage = db.Column(db.Float)
    testimonies_weight = db.Column(db.Float)
    possessions = db.Column(db.Float)
    possessions_percentage = db.Column(db.Float)
    possessions_weight = db.Column(db.Float)
    cic = db.Column(db.Float)
    cic_percentage = db.Column(db.Float)
    cic_weight = db.Column(db.Float)
    targets_fulfillment = db.Column(db.Float)
    incentive_percentage = db.Column(db.Float)
    data_quality = db.Column(db.Float)
    q4_incentive = db.Column(db.Float)
    
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

#####################################################################
#                    API ENDPOINTS                                    #
#####################################################################

@app.route('/api/dashboard', methods=['GET'])
def get_dashboard():
    """Get dashboard overview data"""
    try:
        total_employees = Employee.query.count()
        total_teams = Team.query.count()
        
        # Get recent bonus calculations
        recent_bonuses = db.session.query(BonusCalculation).order_by(BonusCalculation.calculation_date.desc()).limit(5).all()
        
        # Calculate total bonus paid this month
        current_month = datetime.now().month
        current_year = datetime.now().year
        monthly_bonus = db.session.query(db.func.sum(BonusCalculation.bonus_amount)).filter(
            BonusCalculation.month == current_month,
            BonusCalculation.year == current_year
        ).scalar() or 0
        
        dashboard_data = {
            'total_employees': total_employees,
            'total_teams': total_teams,
            'monthly_bonus_total': float(monthly_bonus),
            'recent_calculations': [
                {
                    'id': bonus.id,
                    'employee_name': f"{bonus.employee.name} {bonus.employee.surname}",
                    'bonus_amount': bonus.bonus_amount,
                    'calculation_date': bonus.calculation_date.isoformat()
                } for bonus in recent_bonuses
            ]
        }
        
        return jsonify(dashboard_data), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/teams', methods=['GET'])
def get_teams():
    """Get all teams"""
    try:
        teams = Team.query.all()
        teams_data = []
        
        for team in teams:
            team_data = {
                'id': team.id,
                'name': team.name,
                'description': team.description,
                'employee_count': len(team.employees),
                'created_at': team.created_at.isoformat()
            }
            teams_data.append(team_data)
        
        return jsonify(teams_data), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/teams/<int:team_id>', methods=['GET'])
def get_team(team_id):
    """Get specific team details"""
    try:
        team = Team.query.get_or_404(team_id)
        team_data = {
            'id': team.id,
            'name': team.name,
            'description': team.description,
            'employees': [
                {
                    'id': emp.id,
                    'name': emp.name,
                    'surname': emp.surname,
                    'employee_code': emp.employee_code,
                    'email': emp.email,
                    'position': emp.position,
                    'salary': emp.salary,
                    'target': emp.target
                } for emp in team.employees
            ],
            'incentive_parameters': [
                {
                    'id': param.id,
                    'parameter_name': param.parameter_name,
                    'base_value': param.base_value,
                    'multiplier': param.multiplier,
                    'min_threshold': param.min_threshold,
                    'max_threshold': param.max_threshold,
                    'is_active': param.is_active
                } for param in team.incentive_parameters
            ]
        }
        
        return jsonify(team_data), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/teams/<int:team_id>/download-template', methods=['GET'])
def download_team_template(team_id):
    """Download Excel template for team members"""
    try:
        if not PANDAS_AVAILABLE:
            return jsonify({'error': 'Excel processing not available. Please install pandas and openpyxl.'}), 500
            
        team = Team.query.get_or_404(team_id)
        
        output = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{team.name} Members"
        
        # Define headers and data based on team type
        if team.name.lower() == 'legal team':
            # Legal team template (32 columns)
            headers = [
                'Legal Manager',
                'Employee #',
                'Category',
                'Quarterly Incentive',
                'Team Leader',
                'Lawsuit Presentation (#)',
                'Lawsuit Presentation Target (#)',
                '% Lawsuit Presentation',
                'Lawsuit Presentation Weight',
                'Auction (€)',
                'Auction Target (€)',
                '% Auction',
                'Auction Weight',
                'CDR (€)',
                'CDR Target (€)',
                '% CDR',
                'CDR Weight',
                'Testimonies (€)',
                'Testimonies Target (€)',
                '% Testimonies',
                'Testimonies Weight',
                'Possessions (€)',
                'Possessions Target (€)',
                '% Possessions',
                'Possesions Weight',
                'CIC (€)',
                'CIC Target (€)',
                '% CIC',
                'CIC Weight',
                '% Targets Fulfillment',
                '% Incentive',
                '% Data Quality',
                'TOTAL INCENTIVE %',
                'Q4 TO BE PAID INCENTIVE (80%)'
            ]
            
            # Sample data for Legal team
            sample_data = [
                'John Doe',           # Legal Manager
                'L001',               # Employee #
                'Senior Legal',       # Category
                120000,               # Quarterly Incentive
                'Jane Smith',         # Team Leader
                45,                   # Lawsuit Presentation (#) - from SQL Server
                50,                   # Lawsuit Presentation Target (#)
                '=F2/G2*100',         # % Lawsuit Presentation (calculated)
                20,                   # Lawsuit Presentation Weight
                95000,                # Auction (€) - from SQL Server
                100000,               # Auction Target (€)
                '=J2/K2*100',         # % Auction (calculated)
                25,                   # Auction Weight
                70000,                # CDR (€) - from SQL Server
                75000,                # CDR Target (€)
                '=N2/O2*100',         # % CDR (calculated)
                20,                   # CDR Weight
                22000,                # Testimonies (€) - from SQL Server
                25000,                # Testimonies Target (€)
                '=R2/S2*100',         # % Testimonies (calculated)
                15,                   # Testimonies Weight
                48000,                # Possessions (€) - from SQL Server
                50000,                # Possessions Target (€)
                '=V2/W2*100',         # % Possessions (calculated)
                10,                   # Possesions Weight
                28000,                # CIC (€) - from SQL Server
                30000,                # CIC Target (€)
                '=Z2/AA2*100',        # % CIC (calculated)
                10,                   # CIC Weight
                '=H2*I2/100+J2*L2/100+N2*P2/100+R2*T2/100+V2*X2/100+Z2*AB2/100',  # % Targets Fulfillment (calculated)
                '=AC2*0.85',          # % Incentive (calculated)
                '=RANDBETWEEN(85,100)',  # % Data Quality (calculated)
                '=AD2*AE2/100',       # TOTAL INCENTIVE % (calculated)
                '=AF2*0.8'            # Q4 TO BE PAID INCENTIVE (80%) (calculated)
            ]
            
            # Calculated columns for Legal team (blue background)
            calculated_columns = [8, 12, 16, 20, 24, 28, 30, 31, 32, 33]
            
        else:
            # Servicing team template (18 columns) - existing structure
            headers = [
                'Asset/Sales Manager',
                'Employee Number', 
                'Category',
                'Quarter Incentive Base',
                'Team Leader',
                'Main Portfolio',
                'Cash Flow',
                'Cash Flow Target',
                '% Cash Flow Target',
                'NCF',
                'NCF Target', 
                '% NCF Target',
                'Incentive % CF',
                'NCF Target (Y/N)',
                'Incentive % NCF',
                '% Data Quality',
                'TOTAL INCENTIVE %',
                'Q1 TO BE PAID INCENTIVE (80%)'
            ]
            
            # Sample data for Servicing team
            sample_data = [
                'John Doe',           # Asset/Sales Manager
                'EMP001',             # Employee Number
                'Analyst',            # Category
                50000,                # Quarter Incentive Base
                'Jane Smith',         # Team Leader
                'Portfolio A',        # Main Portfolio
                100000,               # Cash Flow
                120000,               # Cash Flow Target
                '=H2/G2*100',         # % Cash Flow Target (calculated)
                50000,                # NCF
                60000,                # NCF Target
                '=K2/J2*100',         # % NCF Target (calculated)
                '=IF(A2="lezama",IF(I2<60,0,I2),IF(I2<80,0,I2))',  # Incentive % CF (calculated)
                '=IF(K2="","N","Y")', # NCF Target (Y/N) (calculated)
                '=IF(N2="N","N/A",IFERROR(MIN(L2,1),0))',  # Incentive % NCF (calculated)
                '=RANDBETWEEN(85,100)',  # % Data Quality (calculated)
                '=M2+O2',            # TOTAL INCENTIVE % (calculated)
                '=Q2*0.8'            # Q1 TO BE PAID INCENTIVE (80%) (calculated)
            ]
            
            # Calculated columns for Servicing team (blue background)
            calculated_columns = [9, 12, 13, 14, 15, 16, 17, 18]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
        
        # Add sample data row
        for col, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            if col in calculated_columns:  # Calculated columns
                cell.fill = openpyxl.styles.PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                cell.font = openpyxl.styles.Font(italic=True, color="666666")
        
        # Add instructions
        ws.cell(row=4, column=1, value="INSTRUCTIONS:")
        ws.cell(row=4, column=1).font = openpyxl.styles.Font(bold=True)
        
        instructions = [
            "1. Fill in the FIXED columns (white background) with actual data",
            "2. CALCULATED columns (blue background) will be computed automatically",
            "3. Do not modify calculated columns - they will be overwritten",
            "4. Save as .xlsx format before uploading",
            "5. Ensure Employee # is unique for each team member"
        ]
        
        for i, instruction in enumerate(instructions, 5):
            ws.cell(row=i, column=1, value=instruction)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{team.name}_Members_Template.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/teams/<int:team_id>/upload-members', methods=['POST'])
def upload_team_members(team_id):
    """Upload Excel file with team members"""
    try:
        if not PANDAS_AVAILABLE:
            return jsonify({'error': 'Excel processing not available. Please install pandas and openpyxl.'}), 500
            
        team = Team.query.get_or_404(team_id)
        
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith('.xlsx'):
            return jsonify({'error': 'Please upload an Excel file (.xlsx)'}), 400
        
        # Read Excel file
        df = pd.read_excel(file)
        
        # Define required columns based on team type
        if team.name.lower() == 'legal team':
            required_columns = [
                'Legal Manager',
                'Employee #',
                'Category',
                'Quarterly Incentive',
                'Team Leader',
                'Lawsuit Presentation Target (#)',
                'Auction Target (€)',
                'CDR Target (€)',
                'Testimonies Target (€)',
                'Possessions Target (€)',
                'CIC Target (€)'
            ]
        else:
            # Servicing team
            required_columns = [
                'Asset/Sales Manager',
                'Employee Number', 
                'Category',
                'Quarter Incentive Base',
                'Team Leader',
                'Main Portfolio',
                'Cash Flow',
                'Cash Flow Target',
                'NCF',
                'NCF Target'
            ]
        
        # Check for missing required columns
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({'error': f'Missing required columns: {", ".join(missing_columns)}'}), 400
        
        # Process each row and calculate derived fields
        employees_data = []
        for index, row in df.iterrows():
            if team.name.lower() == 'legal team':
                # Process Legal team data
                legal_manager = str(row['Legal Manager']).strip()
                employee_hash = str(row['Employee #']).strip()
                
                # Skip empty rows
                if pd.isna(row['Legal Manager']) or pd.isna(row['Employee #']):
                    continue
                
                # Get target values
                lawsuit_target = float(row['Lawsuit Presentation Target (#)']) if not pd.isna(row['Lawsuit Presentation Target (#)']) else 0
                auction_target = float(row['Auction Target (€)']) if not pd.isna(row['Auction Target (€)']) else 0
                cdr_target = float(row['CDR Target (€)']) if not pd.isna(row['CDR Target (€)']) else 0
                testimonies_target = float(row['Testimonies Target (€)']) if not pd.isna(row['Testimonies Target (€)']) else 0
                possessions_target = float(row['Possessions Target (€)']) if not pd.isna(row['Possessions Target (€)']) else 0
                cic_target = float(row['CIC Target (€)']) if not pd.isna(row['CIC Target (€)']) else 0
                
                # Get actual values from SQL Server for this Legal Manager
                try:
                    # Call our own API endpoint to get SQL Server data
                    import requests
                    quarter = request.args.get('quarter', 'Q4')
                    year = request.args.get('year', '2024')
                    
                    legal_data_response = requests.get(
                        f'http://localhost:5000/api/legal/performance/{legal_manager}',
                        params={'quarter': quarter, 'year': year}
                    )
                    
                    if legal_data_response.status_code == 200:
                        legal_data = legal_data_response.json()
                        lawsuit_actual = legal_data.get('lawsuit_presentation_count', 0)
                        auction_actual = legal_data.get('auction_amount', 0)
                        cdr_actual = legal_data.get('cdr_amount', 0)
                        testimonies_actual = legal_data.get('testimonies_amount', 0)
                        possessions_actual = legal_data.get('possessions_amount', 0)
                        cic_actual = legal_data.get('cic_amount', 0)
                    else:
                        # Fallback to placeholder values if API call fails
                        lawsuit_actual = 45
                        auction_actual = 95000
                        cdr_actual = 70000
                        testimonies_actual = 22000
                        possessions_actual = 48000
                        cic_actual = 28000
                        
                except Exception as e:
                    # Fallback to placeholder values if any error occurs
                    lawsuit_actual = 45
                    auction_actual = 95000
                    cdr_actual = 70000
                    testimonies_actual = 22000
                    possessions_actual = 48000
                    cic_actual = 28000
                
                # Calculate percentages
                lawsuit_percentage = (lawsuit_actual / lawsuit_target * 100) if lawsuit_target > 0 else 0
                auction_percentage = (auction_actual / auction_target * 100) if auction_target > 0 else 0
                cdr_percentage = (cdr_actual / cdr_target * 100) if cdr_target > 0 else 0
                testimonies_percentage = (testimonies_actual / testimonies_target * 100) if testimonies_target > 0 else 0
                possessions_percentage = (possessions_actual / possessions_target * 100) if possessions_target > 0 else 0
                cic_percentage = (cic_actual / cic_target * 100) if cic_target > 0 else 0
                
                # Calculate weighted targets fulfillment
                weights = [20, 25, 20, 15, 10, 10]  # Lawsuit, Auction, CDR, Testimonies, Possessions, CIC
                percentages = [lawsuit_percentage, auction_percentage, cdr_percentage, testimonies_percentage, possessions_percentage, cic_percentage]
                
                targets_fulfillment = sum(p * w / 100 for p, w in zip(percentages, weights))
                incentive_percentage = targets_fulfillment * 0.85  # 85% of targets fulfillment
                data_quality = 95.0  # Default value
                total_incentive = incentive_percentage * data_quality / 100
                q4_incentive = total_incentive * 0.8
                
                employee_data = {
                    'legal_manager': legal_manager,
                    'employee_hash': employee_hash,
                    'category': str(row['Category']).strip() if not pd.isna(row['Category']) else 'Senior Legal',
                    'quarterly_incentive': float(row['Quarterly Incentive']) if not pd.isna(row['Quarterly Incentive']) else 0,
                    'team_leader': str(row['Team Leader']).strip() if not pd.isna(row['Team Leader']) else '',
                    'lawsuit_presentation_target': lawsuit_target,
                    'auction_target': auction_target,
                    'cdr_target': cdr_target,
                    'testimonies_target': testimonies_target,
                    'possessions_target': possessions_target,
                    'cic_target': cic_target,
                    'lawsuit_presentation': lawsuit_actual,
                    'lawsuit_presentation_percentage': round(lawsuit_percentage, 2),
                    'lawsuit_weight': 20,
                    'auction': auction_actual,
                    'auction_percentage': round(auction_percentage, 2),
                    'auction_weight': 25,
                    'cdr': cdr_actual,
                    'cdr_percentage': round(cdr_percentage, 2),
                    'cdr_weight': 20,
                    'testimonies': testimonies_actual,
                    'testimonies_percentage': round(testimonies_percentage, 2),
                    'testimonies_weight': 15,
                    'possessions': possessions_actual,
                    'possessions_percentage': round(possessions_percentage, 2),
                    'possessions_weight': 10,
                    'cic': cic_actual,
                    'cic_percentage': round(cic_percentage, 2),
                    'cic_weight': 10,
                    'targets_fulfillment': round(targets_fulfillment, 2),
                    'incentive_percentage': round(incentive_percentage, 2),
                    'data_quality': data_quality,
                    'total_incentive': round(total_incentive, 2),
                    'q4_incentive': round(q4_incentive, 2),
                    'team_id': team_id,
                    'note': 'Actual values should be calculated from SQL Server data for this Legal Manager'
                }
                
            else:
                # Process Servicing team data (existing logic)
                asset_manager = str(row['Asset/Sales Manager']).strip()
                cash_flow_target = float(row['Cash Flow Target']) if not pd.isna(row['Cash Flow Target']) else 0
                ncf = float(row['NCF']) if not pd.isna(row['NCF']) else 0
                ncf_target = float(row['NCF Target']) if not pd.isna(row['NCF Target']) else 0
                
                # Skip empty rows
                if pd.isna(row['Asset/Sales Manager']) or pd.isna(row['Employee Number']):
                    continue
                
                # TODO: Get actual Cash Flow from SQL Server for this Asset Manager
                # This would require a separate API call to get the aggregated data
                # For now, using the uploaded value as placeholder
                cash_flow = float(row['Cash Flow']) if not pd.isna(row['Cash Flow']) else 0
                
                # Calculate derived fields
                cash_flow_percentage = (cash_flow / cash_flow_target * 100) if cash_flow_target > 0 else 0
                ncf_percentage = (ncf / ncf_target * 100) if ncf_target > 0 else 0
                
                # Calculate incentive percentages based on performance
                incentive_cf = 0
                if asset_manager.lower() == "lezama":
                    # Special case for "lezama" - 60% threshold
                    if cash_flow_percentage >= 60:
                        incentive_cf = cash_flow_percentage
                    else:
                        incentive_cf = 0
                else:
                    # For all other Asset Managers - 80% threshold
                    if cash_flow_percentage >= 80:
                        incentive_cf = cash_flow_percentage
                    else:
                        incentive_cf = 0
                
                # Calculate NCF Target (Y/N) based on whether NCF Target is blank
                ncf_target_yn = 'N' if pd.isna(row['NCF Target']) or row['NCF Target'] == '' else 'Y'
                
                # Calculate Incentive % NCF based on new formula
                if ncf_target_yn == 'N':
                    incentive_ncf = 'N/A'
                else:
                    incentive_ncf = min(ncf_percentage / 100, 1) if ncf_percentage > 0 else 0
                
                # Data quality percentage (placeholder - could be calculated from other metrics)
                data_quality = 95.0  # Default value, could be calculated based on data completeness
                
                # Total incentive percentage
                total_incentive = incentive_cf + (incentive_ncf if incentive_ncf != 'N/A' else 0)
                
                # Q1 incentive to be paid (80% of total)
                q1_incentive = total_incentive * 0.8
                
                employee_data = {
                    'asset_sales_manager': asset_manager,
                    'employee_number': str(row['Employee Number']).strip(),
                    'category': str(row['Category']).strip() if not pd.isna(row['Category']) else 'Analyst',
                    'quarter_incentive_base': float(row['Quarter Incentive Base']) if not pd.isna(row['Quarter Incentive Base']) else 0,
                    'team_leader': str(row['Team Leader']).strip() if not pd.isna(row['Team Leader']) else '',
                    'main_portfolio': str(row['Main Portfolio']).strip() if not pd.isna(row['Main Portfolio']) else '',
                    'cash_flow': cash_flow,
                    'cash_flow_target': cash_flow_target,
                    'cash_flow_percentage': round(cash_flow_percentage, 2),
                    'ncf': ncf,
                    'ncf_target': ncf_target,
                    'ncf_percentage': round(ncf_percentage, 2),
                    'incentive_cf': incentive_cf,
                    'ncf_target_yn': ncf_target_yn,
                    'incentive_ncf': incentive_ncf,
                    'data_quality': data_quality,
                    'total_incentive': round(total_incentive, 2),
                    'q1_incentive': round(q1_incentive, 2),
                    'team_id': team_id,
                    'note': 'Cash Flow should be calculated from SQL Server data using CF + SAA amounts for this Asset Manager'
                }
            
            employees_data.append(employee_data)
        
        return jsonify({
            'message': f'Successfully processed {len(employees_data)} employees',
            'employees': employees_data
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/teams/<int:team_id>/save-members', methods=['POST'])
def save_team_members(team_id):
    """Save uploaded team members to database"""
    try:
        team = Team.query.get_or_404(team_id)
        data = request.get_json()
        
        if 'employees' not in data:
            return jsonify({'error': 'No employees data provided'}), 400
        
        employees_data = data['employees']
        quarter = data.get('quarter', 'Q4')
        year = data.get('year', 2024)
        saved_count = 0
        
        # Clear existing data for this team, quarter, and year
        TeamMemberData.query.filter_by(
            team_id=team_id, 
            quarter=quarter, 
            year=year
        ).delete()
        
        for emp_data in employees_data:
            # Create new team member data record
            team_member = TeamMemberData(
                team_id=team_id,
                quarter=quarter,
                year=year,
                employee_name=emp_data.get('asset_sales_manager') or emp_data.get('legal_manager', ''),
                employee_code=emp_data.get('employee_number') or emp_data.get('employee_hash', ''),
                category=emp_data.get('category', ''),
                team_leader=emp_data.get('team_leader', ''),
                
                # Legal Team fields
                legal_manager=emp_data.get('legal_manager', ''),
                employee_hash=emp_data.get('employee_hash', ''),
                quarterly_incentive=emp_data.get('quarterly_incentive', 0),
                lawsuit_presentation_target=emp_data.get('lawsuit_presentation_target', 0),
                auction_target=emp_data.get('auction_target', 0),
                cdr_target=emp_data.get('cdr_target', 0),
                testimonies_target=emp_data.get('testimonies_target', 0),
                possessions_target=emp_data.get('possessions_target', 0),
                cic_target=emp_data.get('cic_target', 0),
                
                # Servicing Team fields
                asset_sales_manager=emp_data.get('asset_sales_manager', ''),
                employee_number=emp_data.get('employee_number', ''),
                quarter_incentive_base=emp_data.get('quarter_incentive_base', 0),
                main_portfolio=emp_data.get('main_portfolio', ''),
                cash_flow=emp_data.get('cash_flow', 0),
                cash_flow_target=emp_data.get('cash_flow_target', 0),
                ncf=emp_data.get('ncf', 0),
                ncf_target=emp_data.get('ncf_target', 0),
                
                # Calculated fields
                cash_flow_percentage=emp_data.get('cash_flow_percentage', 0),
                ncf_percentage=emp_data.get('ncf_percentage', 0),
                incentive_cf=emp_data.get('incentive_cf', 0),
                total_incentive=emp_data.get('total_incentive', 0),
                q1_incentive=emp_data.get('q1_incentive', 0),
                
                # Legal team calculated fields
                lawsuit_presentation=emp_data.get('lawsuit_presentation', 0),
                lawsuit_presentation_percentage=emp_data.get('lawsuit_presentation_percentage', 0),
                lawsuit_weight=emp_data.get('lawsuit_weight', 0),
                auction=emp_data.get('auction', 0),
                auction_percentage=emp_data.get('auction_percentage', 0),
                auction_weight=emp_data.get('auction_weight', 0),
                cdr=emp_data.get('cdr', 0),
                cdr_percentage=emp_data.get('cdr_percentage', 0),
                cdr_weight=emp_data.get('cdr_weight', 0),
                testimonies=emp_data.get('testimonies', 0),
                testimonies_percentage=emp_data.get('testimonies_percentage', 0),
                testimonies_weight=emp_data.get('testimonies_weight', 0),
                possessions=emp_data.get('possessions', 0),
                possessions_percentage=emp_data.get('possessions_percentage', 0),
                possessions_weight=emp_data.get('possessions_weight', 0),
                cic=emp_data.get('cic', 0),
                cic_percentage=emp_data.get('cic_percentage', 0),
                cic_weight=emp_data.get('cic_weight', 0),
                targets_fulfillment=emp_data.get('targets_fulfillment', 0),
                incentive_percentage=emp_data.get('incentive_percentage', 0),
                data_quality=emp_data.get('data_quality', 0),
                q4_incentive=emp_data.get('q4_incentive', 0)
            )
            
            db.session.add(team_member)
            saved_count += 1
        
        db.session.commit()
        
        return jsonify({
            'message': f'Successfully saved {saved_count} team members to {team.name} for {quarter} {year}',
            'saved_count': saved_count,
            'quarter': quarter,
            'year': year,
            'note': 'Team member data saved to local database. SQL Server remains read-only for performance data.'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/teams/<int:team_id>/members', methods=['GET'])
def get_team_members(team_id):
    """Get saved team members from local database"""
    try:
        team = Team.query.get_or_404(team_id)
        quarter = request.args.get('quarter', 'Q4')
        year = int(request.args.get('year', 2024))
        
        # Get team members from local database
        team_members = TeamMemberData.query.filter_by(
            team_id=team_id,
            quarter=quarter,
            year=year
        ).all()
        
        members_data = []
        for member in team_members:
            member_data = {
                'id': member.id,
                'employee_name': member.employee_name,
                'employee_code': member.employee_code,
                'category': member.category,
                'team_leader': member.team_leader,
                
                # Legal Team fields
                'legal_manager': member.legal_manager,
                'employee_hash': member.employee_hash,
                'quarterly_incentive': member.quarterly_incentive,
                'lawsuit_presentation_target': member.lawsuit_presentation_target,
                'auction_target': member.auction_target,
                'cdr_target': member.cdr_target,
                'testimonies_target': member.testimonies_target,
                'possessions_target': member.possessions_target,
                'cic_target': member.cic_target,
                
                # Servicing Team fields
                'asset_sales_manager': member.asset_sales_manager,
                'employee_number': member.employee_number,
                'quarter_incentive_base': member.quarter_incentive_base,
                'main_portfolio': member.main_portfolio,
                'cash_flow': member.cash_flow,
                'cash_flow_target': member.cash_flow_target,
                'ncf': member.ncf,
                'ncf_target': member.ncf_target,
                
                # Calculated fields
                'cash_flow_percentage': member.cash_flow_percentage,
                'ncf_percentage': member.ncf_percentage,
                'incentive_cf': member.incentive_cf,
                'total_incentive': member.total_incentive,
                'q1_incentive': member.q1_incentive,
                
                # Legal team calculated fields
                'lawsuit_presentation': member.lawsuit_presentation,
                'lawsuit_presentation_percentage': member.lawsuit_presentation_percentage,
                'lawsuit_weight': member.lawsuit_weight,
                'auction': member.auction,
                'auction_percentage': member.auction_percentage,
                'auction_weight': member.auction_weight,
                'cdr': member.cdr,
                'cdr_percentage': member.cdr_percentage,
                'cdr_weight': member.cdr_weight,
                'testimonies': member.testimonies,
                'testimonies_percentage': member.testimonies_percentage,
                'testimonies_weight': member.testimonies_weight,
                'possessions': member.possessions,
                'possessions_percentage': member.possessions_percentage,
                'possessions_weight': member.possessions_weight,
                'cic': member.cic,
                'cic_percentage': member.cic_percentage,
                'cic_weight': member.cic_weight,
                'targets_fulfillment': member.targets_fulfillment,
                'incentive_percentage': member.incentive_percentage,
                'data_quality': member.data_quality,
                'q4_incentive': member.q4_incentive,
                
                'created_at': member.created_at.isoformat() if member.created_at else None,
                'updated_at': member.updated_at.isoformat() if member.updated_at else None
            }
            members_data.append(member_data)
        
        return jsonify({
            'team_id': team_id,
            'team_name': team.name,
            'quarter': quarter,
            'year': year,
            'members': members_data,
            'count': len(members_data)
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/performance', methods=['GET'])
def get_performance_data():
    """Get performance data"""
    try:
        performance_records = PerformanceRecord.query.all()
        performance_data = []
        
        for record in performance_records:
            record_data = {
                'id': record.id,
                'employee_name': f"{record.employee.name} {record.employee.surname}",
                'team_name': record.employee.team.name,
                'month': record.month,
                'year': record.year,
                'productivity_score': record.productivity_score,
                'quality_score': record.quality_score,
                'attendance_score': record.attendance_score,
                'overall_score': record.overall_score
            }
            performance_data.append(record_data)
        
        return jsonify(performance_data), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/performance/team/<int:team_id>', methods=['GET'])
def get_team_performance(team_id):
    """Get performance data for a specific team"""
    try:
        team = Team.query.get_or_404(team_id)
        
        # Placeholder queries for each team - you can replace these with actual SQL queries
        if team_id == 1:  # Legal Team
            # PLACEHOLDER QUERY FOR LEGAL TEAM
            # SELECT 
            #   COUNT(DISTINCT employee_id) as total_employees,
            #   AVG(productivity_score) as avg_productivity,
            #   AVG(quality_score) as avg_quality,
            #   AVG(attendance_score) as avg_attendance,
            #   AVG(overall_score) as avg_overall
            # FROM performance_records pr
            # JOIN employees e ON pr.employee_id = e.id
            # WHERE e.team_id = 1 AND pr.year = 2024 AND pr.month IN (10, 11, 12)
            
            performance_data = {
                'team_id': team_id,
                'team_name': team.name,
                'quarter': 'Q4 2024',
                'total_employees': 8,
                'avg_productivity': 87,
                'avg_quality': 92,
                'avg_attendance': 95,
                'avg_overall': 91,
                'top_performers': [
                    {'employee_name': 'Alice Smith', 'overall_score': 96},
                    {'employee_name': 'Bob Johnson', 'overall_score': 94},
                    {'employee_name': 'Carol White', 'overall_score': 92}
                ],
                'quarterly_trend': [
                    {'month': 'Oct', 'avg_score': 85},
                    {'month': 'Nov', 'avg_score': 88},
                    {'month': 'Dec', 'avg_score': 91}
                ],
                'performance_distribution': [
                    {'range': '90-100%', 'count': 4},
                    {'range': '80-89%', 'count': 3},
                    {'range': '70-79%', 'count': 1},
                    {'range': '60-69%', 'count': 0}
                ]
            }
            
        elif team_id == 2:  # Loan Team
            # PLACEHOLDER QUERY FOR LOAN TEAM
            # SELECT 
            #   COUNT(DISTINCT employee_id) as total_employees,
            #   AVG(productivity_score) as avg_productivity,
            #   AVG(quality_score) as avg_quality,
            #   AVG(attendance_score) as avg_attendance,
            #   AVG(overall_score) as avg_overall
            # FROM performance_records pr
            # JOIN employees e ON pr.employee_id = e.id
            # WHERE e.team_id = 2 AND pr.year = 2024 AND pr.month IN (10, 11, 12)
            
            performance_data = {
                'team_id': team_id,
                'team_name': team.name,
                'quarter': 'Q4 2024',
                'total_employees': 12,
                'avg_productivity': 89,
                'avg_quality': 88,
                'avg_attendance': 93,
                'avg_overall': 90,
                'top_performers': [
                    {'employee_name': 'David Brown', 'overall_score': 95},
                    {'employee_name': 'Eve Davis', 'overall_score': 93},
                    {'employee_name': 'Frank Wilson', 'overall_score': 91}
                ],
                'quarterly_trend': [
                    {'month': 'Oct', 'avg_score': 87},
                    {'month': 'Nov', 'avg_score': 89},
                    {'month': 'Dec', 'avg_score': 90}
                ],
                'performance_distribution': [
                    {'range': '90-100%', 'count': 6},
                    {'range': '80-89%', 'count': 4},
                    {'range': '70-79%', 'count': 2},
                    {'range': '60-69%', 'count': 0}
                ]
            }
            
        elif team_id == 3:  # Servicing Team
            # PLACEHOLDER QUERY FOR SERVICING TEAM
            # SELECT 
            #   COUNT(DISTINCT employee_id) as total_employees,
            #   AVG(productivity_score) as avg_productivity,
            #   AVG(quality_score) as avg_quality,
            #   AVG(attendance_score) as avg_attendance,
            #   AVG(overall_score) as avg_overall
            # FROM performance_records pr
            # JOIN employees e ON pr.employee_id = e.id
            # WHERE e.team_id = 3 AND pr.year = 2024 AND pr.month IN (10, 11, 12)
            
            performance_data = {
                'team_id': team_id,
                'team_name': team.name,
                'quarter': 'Q4 2024',
                'total_employees': 15,
                'avg_productivity': 85,
                'avg_quality': 86,
                'avg_attendance': 91,
                'avg_overall': 87,
                'top_performers': [
                    {'employee_name': 'Grace Lee', 'overall_score': 93},
                    {'employee_name': 'Henry Chen', 'overall_score': 91},
                    {'employee_name': 'Ivy Taylor', 'overall_score': 89}
                ],
                'quarterly_trend': [
                    {'month': 'Oct', 'avg_score': 84},
                    {'month': 'Nov', 'avg_score': 86},
                    {'month': 'Dec', 'avg_score': 87}
                ],
                'performance_distribution': [
                    {'range': '90-100%', 'count': 5},
                    {'range': '80-89%', 'count': 7},
                    {'range': '70-79%', 'count': 3},
                    {'range': '60-69%', 'count': 0}
                ]
            }
            
        else:
            return jsonify({'error': 'Team not found'}), 404
        
        return jsonify(performance_data), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/performance/team/<int:team_id>/refresh', methods=['GET'])
def refresh_team_performance(team_id):
    """Refresh performance data for a specific team from external SQL server"""
    try:
        team = Team.query.get_or_404(team_id)
        
        # Get quarter parameter from request (default to Q4 2024)
        quarter = request.args.get('quarter', 'Q4')
        year = request.args.get('year', '2024')
        
        # SQL Server Connection Placeholder - Uncomment and configure when ready
        # import pyodbc
        # try:
        #     connection = pyodbc.connect(SQL_SERVER_CONNECTION_STRING)
        #     cursor = connection.cursor()
        #     
        #     if team_id == 1:  # Legal Team - Actual Query
        #         # Get quarter date range
        #         quarter_range = QUARTER_DATES.get(quarter, QUARTER_DATES['Q4'])
        #         start_date = f"{year}-{quarter_range['start']}"
        #         end_date = f"{year}-{quarter_range['end']}"
        #         
        #         # Execute the actual Legal Team query
        #         query = LEGAL_TEAM_QUERY.format(start_date=start_date, end_date=end_date)
        #         cursor.execute(query)
        #         results = cursor.fetchall()
        #         
        #         # Process the results
        #         total_acts = len(results)
        #         total_amount = sum(row[10] for row in results if row[10] is not None)
        #         
        #         # Calculate metrics by bucket and lawyer
        #         bucket_distribution = {}
        #         lawyer_distribution = {}
        #         legal_stages = {}
        #         
        #         for row in results:
        #             lawyer = row[0] if row[0] else 'Unknown'
        #             bucket = row[13] if row[13] else 'Unknown'
        #             legal_stage = row[1] if row[1] else 'Unknown'
        #             amount = row[10] if row[10] else 0
        #             
        #             bucket_distribution[bucket] = bucket_distribution.get(bucket, 0) + 1
        #             lawyer_distribution[lawyer] = lawyer_distribution.get(lawyer, 0) + 1
        #             legal_stages[legal_stage] = legal_stages.get(legal_stage, 0) + 1
        #         
        #         # Get top lawyers by number of acts
        #         top_lawyers = sorted(lawyer_distribution.items(), key=lambda x: x[1], reverse=True)[:3]
        #         
        #         performance_data = {
        #             'team_id': team_id,
        #             'team_name': team.name,
        #             'quarter': f'{quarter} {year}',
        #             'total_legal_acts': total_acts,
        #             'total_amount': round(total_amount, 2),
        #             'avg_amount_per_act': round(total_amount / total_acts, 2) if total_acts > 0 else 0,
        #             'bucket_distribution': bucket_distribution,
        #             'lawyer_distribution': lawyer_distribution,
        #             'legal_stages': legal_stages,
        #             'data_source': 'SQL Server - LegalActActivity',
        #             'query_period': f'{start_date} to {end_date}',
        #             'country_id': 2
        #         }
        #         
        #     elif team_id == 3:  # Servicing Team - Actual Query
        #         # Get quarter date range
        #         quarter_range = QUARTER_DATES.get(quarter, QUARTER_DATES['Q4'])
        #         start_date = f"{year}-{quarter_range['start']}"
        #         end_date = f"{year}-{quarter_range['end']}"
        #         
        #         # Execute the actual Servicing Team query
        #         query = SERVICING_TEAM_QUERY.format(start_date=start_date, end_date=end_date)
        #         cursor.execute(query)
        #         results = cursor.fetchall()
        #         
        #         # Process the results
        #         total_collections = len(results)
        #         total_amount = sum(row[1] for row in results if row[1] is not None)
        #         
        #         # Calculate metrics by category
        #         category_groups = {}
        #         cf_types = {}
        #         
        #         for row in results:
        #             category_group = row[4] if row[4] else 'Unknown'
        #             cf_type = row[5] if row[5] else 'Unknown'
        #             amount = row[1] if row[1] else 0
        #             
        #             category_groups[category_group] = category_groups.get(category_group, 0) + 1
        #             cf_types[cf_type] = cf_types.get(cf_type, 0) + 1
        #         
        #         performance_data = {
        #             'team_id': team_id,
        #             'team_name': team.name,
        #             'quarter': f'{quarter} {year}',
        #             'total_collections': total_collections,
        #             'total_amount': round(total_amount, 2),
        #             'avg_amount_per_collection': round(total_amount / total_collections, 2) if total_collections > 0 else 0,
        #             'category_distribution': category_groups,
        #             'cf_type_distribution': cf_types,
        #             'data_source': 'SQL Server - CollectionDetail',
        #             'query_period': f'{start_date} to {end_date}',
        #             'country': 'ESPANA'
        #         }
        #         
        #     elif team_id == 2:  # Loan Team - Placeholder
        #         # Similar structure for Loan Team
        #         performance_data = {
        #             'team_id': team_id,
        #             'team_name': team.name,
        #             'quarter': f'{quarter} {year}',
        #             'data_source': 'SQL Server - Placeholder'
        #         }
        #     
        #     cursor.close()
        #     connection.close()
        #     
        # except Exception as e:
        #     return jsonify({'error': f'SQL Server connection failed: {str(e)}'}), 500
        
        # For now, return updated mock data (replace this section with actual SQL Server queries above)
        if team_id == 1:  # Legal Team - Mock data with actual query structure
            # Get quarter date range for display
            quarter_range = QUARTER_DATES.get(quarter, QUARTER_DATES['Q4'])
            start_date = f"{year}-{quarter_range['start']}"
            end_date = f"{year}-{quarter_range['end']}"
            
            performance_data = {
                'team_id': team_id,
                'team_name': team.name,
                'quarter': f'{quarter} {year}',
                'total_legal_acts': 850,  # Mock data
                'total_amount': 12500000.75,  # Mock data
                'avg_amount_per_act': 14705.88,  # Mock data
                'bucket_distribution': {
                    'Auction': 180,
                    'Assigment of awarding': 120,
                    'Cash In Court': 200,
                    'Testimony': 150,
                    'Demands': 100,
                    'Possession': 100
                },
                'lawyer_distribution': {
                    'Alice Smith': 45,
                    'Bob Johnson': 42,
                    'Carol White': 38,
                    'David Brown': 35,
                    'Eve Davis': 32
                },
                'legal_stages': {
                    'LAWSUIT ADMISSION': 200,
                    'JUDICIAL POSSESSION': 150,
                    'AUCTION PROCESS': 180,
                    'AWARDING': 120,
                    'EXECUTION': 200
                },
                'data_source': 'Mock Data (Replace with SQL Server - LegalActActivity)',
                'query_period': f'{start_date} to {end_date}',
                'country_id': 2,
                'top_performers': [
                    {'employee_name': 'Alice Smith', 'legal_acts': 45},
                    {'employee_name': 'Bob Johnson', 'legal_acts': 42},
                    {'employee_name': 'Carol White', 'legal_acts': 38}
                ],
                'quarterly_trend': [
                    {'month': 'Oct', 'legal_acts': 280},
                    {'month': 'Nov', 'legal_acts': 290},
                    {'month': 'Dec', 'legal_acts': 280}
                ]
            }
            
        elif team_id == 2:  # Loan Team
            performance_data = {
                'team_id': team_id,
                'team_name': team.name,
                'quarter': f'{quarter} {year}',
                'total_employees': 12,
                'avg_productivity': 89,
                'avg_quality': 88,
                'avg_attendance': 93,
                'avg_overall': 90,
                'data_source': 'Mock Data (Replace with SQL Server)',
                'top_performers': [
                    {'employee_name': 'David Brown', 'overall_score': 95},
                    {'employee_name': 'Eve Davis', 'overall_score': 93},
                    {'employee_name': 'Frank Wilson', 'overall_score': 91}
                ],
                'quarterly_trend': [
                    {'month': 'Oct', 'avg_score': 87},
                    {'month': 'Nov', 'avg_score': 89},
                    {'month': 'Dec', 'avg_score': 90}
                ],
                'performance_distribution': [
                    {'range': '90-100%', 'count': 6},
                    {'range': '80-89%', 'count': 4},
                    {'range': '70-79%', 'count': 2},
                    {'range': '60-69%', 'count': 0}
                ]
            }
            
        elif team_id == 3:  # Servicing Team - Mock data with actual query structure
            # Get quarter date range for display
            quarter_range = QUARTER_DATES.get(quarter, QUARTER_DATES['Q4'])
            start_date = f"{year}-{quarter_range['start']}"
            end_date = f"{year}-{quarter_range['end']}"
            
            performance_data = {
                'team_id': team_id,
                'team_name': team.name,
                'quarter': f'{quarter} {year}',
                'total_collections': 1250,  # Mock data
                'total_amount': 2850000.50,  # Mock data
                'avg_amount_per_collection': 2280.00,  # Mock data
                'category_distribution': {
                    'Amicable Debtor Resolution': 450,
                    'Real Estate': 380,
                    'Third Part Resolution': 420
                },
                'cf_type_distribution': {
                    'CF': 680,
                    'Legal': 320,
                    'SSA': 180,
                    'Non-CF': 70
                },
                'data_source': 'Mock Data (Replace with SQL Server - CollectionDetail)',
                'query_period': f'{start_date} to {end_date}',
                'country': 'ESPANA',
                'top_performers': [
                    {'employee_name': 'Grace Lee', 'collections': 45},
                    {'employee_name': 'Henry Chen', 'collections': 42},
                    {'employee_name': 'Ivy Taylor', 'collections': 38}
                ],
                'quarterly_trend': [
                    {'month': 'Oct', 'collections': 400},
                    {'month': 'Nov', 'collections': 420},
                    {'month': 'Dec', 'collections': 430}
                ]
            }
            
        else:
            return jsonify({'error': 'Team not found'}), 404
        
        return jsonify(performance_data), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/incentives', methods=['GET'])
def get_incentive_parameters():
    """Get all incentive parameters"""
    try:
        parameters = IncentiveParameter.query.all()
        parameters_data = []
        
        for param in parameters:
            param_data = {
                'id': param.id,
                'team_id': param.team_id,
                'team_name': param.team.name,
                'category': param.category,
                'base_bonus': param.base_bonus,
                'parameter_name': param.parameter_name,
                'base_value': param.base_value,
                'multiplier': param.multiplier,
                'min_threshold': param.min_threshold,
                'max_threshold': param.max_threshold,
                'is_active': param.is_active
            }
            parameters_data.append(param_data)
        
        return jsonify(parameters_data), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/incentives', methods=['POST'])
def create_incentive_parameter():
    """Create a new incentive parameter"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['team_id', 'category', 'base_bonus', 'quarter', 'year']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Check if team exists
        team = Team.query.get(data['team_id'])
        if not team:
            return jsonify({'error': 'Team not found'}), 404
        
        # Create new parameter
        new_parameter = IncentiveParameter(
            team_id=data['team_id'],
            category=data['category'],
            base_bonus=data.get('base_bonus', 0.0),
            parameter_name=f"{data['category']}_bonus",  # Auto-generate parameter name
            base_value=data.get('base_value', 0.0),
            multiplier=data.get('multiplier', 1.0),
            min_threshold=data.get('min_threshold', 0.0),
            max_threshold=data.get('max_threshold', 100.0),
            is_active=data.get('is_active', True),
            quarter=data['quarter'],
            year=data['year']
        )
        
        db.session.add(new_parameter)
        db.session.commit()
        
        # Return the created parameter
        param_data = {
            'id': new_parameter.id,
            'team_id': new_parameter.team_id,
            'team_name': new_parameter.team.name,
            'category': new_parameter.category,
            'base_bonus': new_parameter.base_bonus,
            'parameter_name': new_parameter.parameter_name,
            'base_value': new_parameter.base_value,
            'multiplier': new_parameter.multiplier,
            'min_threshold': new_parameter.min_threshold,
            'max_threshold': new_parameter.max_threshold,
            'is_active': new_parameter.is_active,
            'quarter': new_parameter.quarter,
            'year': new_parameter.year
        }
        
        return jsonify(param_data), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/incentives/<int:param_id>', methods=['PUT'])
def update_incentive_parameter(param_id):
    """Update an existing incentive parameter"""
    try:
        parameter = IncentiveParameter.query.get_or_404(param_id)
        data = request.get_json()
        
        # Update fields if provided
        if 'category' in data:
            parameter.category = data['category']
        if 'base_bonus' in data:
            parameter.base_bonus = data['base_bonus']
        if 'base_value' in data:
            parameter.base_value = data['base_value']
        if 'multiplier' in data:
            parameter.multiplier = data['multiplier']
        if 'min_threshold' in data:
            parameter.min_threshold = data['min_threshold']
        if 'max_threshold' in data:
            parameter.max_threshold = data['max_threshold']
        if 'is_active' in data:
            parameter.is_active = data['is_active']
        
        db.session.commit()
        
        # Return the updated parameter
        param_data = {
            'id': parameter.id,
            'team_id': parameter.team_id,
            'team_name': parameter.team.name,
            'category': parameter.category,
            'base_bonus': parameter.base_bonus,
            'parameter_name': parameter.parameter_name,
            'base_value': parameter.base_value,
            'multiplier': parameter.multiplier,
            'min_threshold': parameter.min_threshold,
            'max_threshold': parameter.max_threshold,
            'is_active': parameter.is_active
        }
        
        return jsonify(param_data), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/incentives/<int:param_id>', methods=['DELETE'])
def delete_incentive_parameter(param_id):
    """Delete an incentive parameter"""
    try:
        parameter = IncentiveParameter.query.get_or_404(param_id)
        db.session.delete(parameter)
        db.session.commit()
        
        return jsonify({'message': 'Parameter deleted successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/calculate-bonus', methods=['POST'])
def calculate_bonus():
    """Calculate bonus for an employee"""
    try:
        data = request.get_json()
        employee_id = data.get('employee_id')
        month = data.get('month', datetime.now().month)
        year = data.get('year', datetime.now().year)
        quarter = f"Q{(month - 1) // 3 + 1}"  # Calculate quarter from month
        
        employee = Employee.query.get_or_404(employee_id)
        performance = PerformanceRecord.query.filter_by(
            employee_id=employee_id,
            month=month,
            year=year
        ).first()
        
        if not performance:
            return jsonify({'error': 'Performance record not found'}), 404
        
        # Get team incentive parameters for the specific quarter and year
        team_params = IncentiveParameter.query.filter_by(
            team_id=employee.team_id,
            is_active=True,
            quarter=quarter,
            year=year
        ).all()
        
        if not team_params:
            return jsonify({'error': f'No incentive parameters found for {quarter} {year}'}), 404
        
        # Calculate bonus based on performance and parameters
        base_bonus = employee.salary * 0.1  # 10% base bonus
        performance_multiplier = performance.overall_score / 100
        
        # Apply team-specific parameters
        for param in team_params:
            if param.parameter_name == 'performance_multiplier':
                performance_multiplier *= param.multiplier
            elif param.parameter_name == 'base_bonus_percentage':
                base_bonus = employee.salary * (param.base_value / 100)
        
        final_bonus = base_bonus * performance_multiplier
        
        # Save calculation with quarter information
        bonus_calc = BonusCalculation(
            employee_id=employee_id,
            month=month,
            year=year,
            quarter=quarter,
            base_salary=employee.salary,
            performance_score=performance.overall_score,
            bonus_amount=final_bonus
        )
        
        db.session.add(bonus_calc)
        db.session.commit()
        
        return jsonify({
            'employee_id': employee_id,
            'quarter': quarter,
            'year': year,
            'base_salary': employee.salary,
            'performance_score': performance.overall_score,
            'bonus_amount': final_bonus
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/seed-data', methods=['POST'])
def seed_data():
    """Seed initial data for development"""
    try:
        # Create teams
        teams_data = [
            {'name': 'Legal Team', 'description': 'Legal and compliance team'},
            {'name': 'Loan Team', 'description': 'Loan processing and underwriting team'},
            {'name': 'Servicing Team', 'description': 'Customer service and loan servicing team'}
        ]
        
        for team_data in teams_data:
            team = Team(**team_data)
            db.session.add(team)
        
        db.session.commit()
        
        # Get created teams
        legal_team = Team.query.filter_by(name='Legal Team').first()
        loan_team = Team.query.filter_by(name='Loan Team').first()
        servicing_team = Team.query.filter_by(name='Servicing Team').first()
        
        # Create employees with new fields
        employees_data = [
            {'name': 'John', 'surname': 'Smith', 'employee_code': 'EMP001', 'email': 'john.smith@company.com', 'position': 'Senior Legal Counsel', 'salary': 120000, 'target': 100000, 'team_id': legal_team.id},
            {'name': 'Sarah', 'surname': 'Johnson', 'employee_code': 'EMP002', 'email': 'sarah.johnson@company.com', 'position': 'Legal Assistant', 'salary': 65000, 'target': 50000, 'team_id': legal_team.id},
            {'name': 'Mike', 'surname': 'Davis', 'employee_code': 'EMP003', 'email': 'mike.davis@company.com', 'position': 'Loan Officer', 'salary': 85000, 'target': 75000, 'team_id': loan_team.id},
            {'name': 'Lisa', 'surname': 'Wilson', 'employee_code': 'EMP004', 'email': 'lisa.wilson@company.com', 'position': 'Underwriter', 'salary': 95000, 'target': 85000, 'team_id': loan_team.id},
            {'name': 'Tom', 'surname': 'Brown', 'employee_code': 'EMP005', 'email': 'tom.brown@company.com', 'position': 'Customer Service Manager', 'salary': 75000, 'target': 60000, 'team_id': servicing_team.id},
            {'name': 'Emily', 'surname': 'Chen', 'employee_code': 'EMP006', 'email': 'emily.chen@company.com', 'position': 'Service Representative', 'salary': 55000, 'target': 45000, 'team_id': servicing_team.id}
        ]
        
        for emp_data in employees_data:
            employee = Employee(**emp_data)
            db.session.add(employee)
        
        db.session.commit()
        
        # Create incentive parameters
        incentive_params = [
            {'team_id': legal_team.id, 'parameter_name': 'performance_multiplier', 'base_value': 1.0, 'multiplier': 1.2, 'min_threshold': 0, 'max_threshold': 100},
            {'team_id': loan_team.id, 'parameter_name': 'performance_multiplier', 'base_value': 1.0, 'multiplier': 1.3, 'min_threshold': 0, 'max_threshold': 100},
            {'team_id': servicing_team.id, 'parameter_name': 'performance_multiplier', 'base_value': 1.0, 'multiplier': 1.1, 'min_threshold': 0, 'max_threshold': 100}
        ]
        
        for param_data in incentive_params:
            param = IncentiveParameter(**param_data)
            db.session.add(param)
        
        db.session.commit()
        
        return jsonify({'message': 'Data seeded successfully'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/servicing/cash-flow/<asset_manager>', methods=['GET'])
def get_asset_manager_cash_flow(asset_manager):
    """Get aggregated Cash Flow data for a specific Asset Manager from SQL Server"""
    try:
        # Get quarter parameter from request (default to Q4 2024)
        quarter = request.args.get('quarter', 'Q4')
        year = request.args.get('year', '2024')
        
        # Get quarter date range
        quarter_range = QUARTER_DATES.get(quarter, QUARTER_DATES['Q4'])
        start_date = f"{year}-{quarter_range['start']}"
        end_date = f"{year}-{quarter_range['end']}"
        
        # SQL Server Connection Placeholder - Uncomment and configure when ready
        # import pyodbc
        # try:
        #     connection = pyodbc.connect(SQL_SERVER_CONNECTION_STRING)
        #     cursor = connection.cursor()
        #     
        #     # Execute the aggregated query for the specific Asset Manager
        #     query = f"""
        #     SELECT 
        #         AssetManager,
        #         COUNT(*) as total_collections,
        #         SUM(TotalAmount) as total_amount,
        #         SUM(CASE WHEN CFType IN ('CF', 'SSA') THEN TotalAmount ELSE 0 END) as cash_flow_amount,
        #         SUM(CASE WHEN CFType = 'CF' THEN TotalAmount ELSE 0 END) as cf_amount,
        #         SUM(CASE WHEN CFType = 'SSA' THEN TotalAmount ELSE 0 END) as ssa_amount,
        #         COUNT(CASE WHEN CFType IN ('CF', 'SSA') THEN 1 END) as cf_collections_count,
        #         -- NCF calculation: count Non-CF collections where flag column = 0
        #         COUNT(CASE WHEN CFType = 'Non-CF' AND FlagColumn = 0 THEN 1 END) as ncf_count
        #     FROM (
        #         SELECT 
        #             TotalAmount,
        #             FlagColumn,  -- Add the flag column from your SQL Server table
        #             CASE 
        #                 WHEN CollectionCategory IN (
        #                     'Assignment Of Award – Sale Third Party', 
        #                     'Cash In Court Third Party - Sale At Auction', 
        #                     'Cash In Court Third Party - Servicing'
        #                 ) THEN 'SSA'
        #                 WHEN CollectionCategory IN (
        #                     'Rent', 'Pspa', 'Sale Deed', 'Workout', 'Prepayment Partial', 
        #                     'Prepayment Full', 'Loan Sale', 'Installment', 
        #                     'Discounted Payoff  - Secured', 'Discounted Payoff  - Unsecured', 
        #                     'Collateral Sale'
        #                 ) THEN 'CF'
        #                 WHEN CollectionCategory IN (
        #                     'Cash In Court', 'Cash In Court Third Party - Secured', 
        #                     'Cash In Court Third Party - Unsecured'
        #                 ) THEN 'Legal'
        #                 ELSE 'Non-CF'
        #             END AS CFType,
        #             AssetManager
        #         FROM CollectionDetail
        #         WHERE Country = 'ESPANA'
        #         AND ReceivedDate BETWEEN '{start_date}' AND '{end_date}'
        #         AND AssetManager = ?
        #     ) AS subquery
        #     GROUP BY AssetManager
        #     """
        #     
        #     cursor.execute(query, (asset_manager,))
        #     result = cursor.fetchone()
        #     
        #     if result:
        #         cash_flow_data = {
        #             'asset_manager': result[0],
        #             'total_collections': result[1],
        #             'total_amount': round(result[2], 2) if result[2] else 0,
        #             'cash_flow_amount': round(result[3], 2) if result[3] else 0,  # CF + SAA
        #             'cf_amount': round(result[4], 2) if result[4] else 0,
        #             'ssa_amount': round(result[5], 2) if result[5] else 0,
        #             'cf_collections_count': result[6],
        #             'ncf_count': result[7],  # NCF count
        #             'quarter': f'{quarter} {year}',
        #             'query_period': f'{start_date} to {end_date}'
        #         }
        #     else:
        #         cash_flow_data = {
        #             'asset_manager': asset_manager,
        #             'total_collections': 0,
        #             'total_amount': 0,
        #             'cash_flow_amount': 0,
        #             'cf_amount': 0,
        #             'ssa_amount': 0,
        #             'cf_collections_count': 0,
        #             'ncf_count': 0,
        #             'quarter': f'{quarter} {year}',
        #             'query_period': f'{start_date} to {end_date}',
        #             'note': 'No data found for this Asset Manager'
        #         }
        #     
        #     cursor.close()
        #     connection.close()
        #     
        # except Exception as e:
        #     return jsonify({'error': f'SQL Server connection failed: {str(e)}'}), 500
        
        # For now, return mock data (replace with actual SQL Server query above)
        cash_flow_data = {
            'asset_manager': asset_manager,
            'total_collections': 150,
            'total_amount': 2500000.00,
            'cash_flow_amount': 1800000.00,  # CF + SAA amounts
            'cf_amount': 1200000.00,
            'ssa_amount': 600000.00,
            'cf_collections_count': 85,
            'ncf_count': 25,  # NCF count (Non-CF collections where flag = 0)
            'quarter': f'{quarter} {year}',
            'query_period': f'{start_date} to {end_date}',
            'note': 'Mock data - replace with actual SQL Server query'
        }
        
        return jsonify(cash_flow_data), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/legal/performance/<legal_manager>', methods=['GET'])
def get_legal_manager_performance(legal_manager):
    """Get aggregated Legal performance data for a specific Legal Manager from SQL Server"""
    try:
        # Get quarter parameter from request (default to Q4 2024)
        quarter = request.args.get('quarter', 'Q4')
        year = request.args.get('year', '2024')
        
        # Get quarter date range
        quarter_range = QUARTER_DATES.get(quarter, QUARTER_DATES['Q4'])
        start_date = f"{year}-{quarter_range['start']}"
        end_date = f"{year}-{quarter_range['end']}"
        
        # SQL Server Connection Placeholder - Uncomment and configure when ready
        # import pyodbc
        # try:
        #     connection = pyodbc.connect(SQL_SERVER_CONNECTION_STRING)
        #     cursor = connection.cursor()
        #     
        #     # Execute the aggregated query for the specific Legal Manager
        #     query = f"""
        #     SELECT 
        #         l.InternalLawyerName,
        #         -- Lawsuit Presentation count
        #         COUNT(CASE WHEN Bucket = 'Demands' THEN 1 END) as lawsuit_presentation_count,
        #         -- Auction amount
        #         SUM(CASE WHEN Bucket = 'Auction' THEN act.ActAmount ELSE 0 END) as auction_amount,
        #         -- CDR amount (Assigment of awarding)
        #         SUM(CASE WHEN Bucket = 'Assigment of awarding' THEN act.ActAmount ELSE 0 END) as cdr_amount,
        #         -- Testimonies amount
        #         SUM(CASE WHEN Bucket = 'Testimony' THEN act.ActAmount ELSE 0 END) as testimonies_amount,
        #         -- Possessions amount
        #         SUM(CASE WHEN Bucket = 'Possession' THEN act.ActAmount ELSE 0 END) as possessions_amount,
        #         -- CIC amount (Cash In Court)
        #         SUM(CASE WHEN Bucket = 'Cash In Court' THEN act.ActAmount ELSE 0 END) as cic_amount,
        #         -- Total legal acts
        #         COUNT(*) as total_legal_acts
        #     FROM (
        #         {LEGAL_TEAM_QUERY.format(start_date=start_date, end_date=end_date)}
        #     ) AS legal_data
        #     WHERE l.InternalLawyerName = ?
        #     GROUP BY l.InternalLawyerName
        #     """
        #     
        #     cursor.execute(query, (legal_manager,))
        #     result = cursor.fetchone()
        #     
        #     if result:
        #         legal_performance_data = {
        #             'legal_manager': result[0],
        #             'lawsuit_presentation_count': result[1] or 0,
        #             'auction_amount': round(result[2], 2) if result[2] else 0,
        #             'cdr_amount': round(result[3], 2) if result[3] else 0,
        #             'testimonies_amount': round(result[4], 2) if result[4] else 0,
        #             'possessions_amount': round(result[5], 2) if result[5] else 0,
        #             'cic_amount': round(result[6], 2) if result[6] else 0,
        #             'total_legal_acts': result[7] or 0,
        #             'quarter': f'{quarter} {year}',
        #             'query_period': f'{start_date} to {end_date}'
        #         }
        #     else:
        #         legal_performance_data = {
        #             'legal_manager': legal_manager,
        #             'lawsuit_presentation_count': 0,
        #             'auction_amount': 0,
        #             'cdr_amount': 0,
        #             'testimonies_amount': 0,
        #             'possessions_amount': 0,
        #             'cic_amount': 0,
        #             'total_legal_acts': 0,
        #             'quarter': f'{quarter} {year}',
        #             'query_period': f'{start_date} to {end_date}',
        #             'note': 'No data found for this Legal Manager'
        #         }
        #     
        #     cursor.close()
        #     connection.close()
        #     
        # except Exception as e:
        #     return jsonify({'error': f'SQL Server connection failed: {str(e)}'}), 500
        
        # For now, return mock data (replace with actual SQL Server query above)
        legal_performance_data = {
            'legal_manager': legal_manager,
            'lawsuit_presentation_count': 45,  # From SQL Server: COUNT WHERE Bucket = 'Demands'
            'auction_amount': 95000.00,        # From SQL Server: SUM WHERE Bucket = 'Auction'
            'cdr_amount': 70000.00,            # From SQL Server: SUM WHERE Bucket = 'Assigment of awarding'
            'testimonies_amount': 22000.00,    # From SQL Server: SUM WHERE Bucket = 'Testimony'
            'possessions_amount': 48000.00,    # From SQL Server: SUM WHERE Bucket = 'Possession'
            'cic_amount': 28000.00,            # From SQL Server: SUM WHERE Bucket = 'Cash In Court'
            'total_legal_acts': 150,           # From SQL Server: COUNT(*)
            'quarter': f'{quarter} {year}',
            'query_period': f'{start_date} to {end_date}',
            'note': 'Mock data - replace with actual SQL Server query',
            'sql_query_used': LEGAL_TEAM_QUERY.format(start_date=start_date, end_date=end_date)
        }
        
        return jsonify(legal_performance_data), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/test-db', methods=['GET'])
def test_database():
    """Test database connection and models"""
    try:
        # Test if we can query the database
        team_count = Team.query.count()
        employee_count = Employee.query.count()
        team_member_data_count = TeamMemberData.query.count()
        
        return jsonify({
            'status': 'success',
            'message': 'Database connection successful',
            'team_count': team_count,
            'employee_count': employee_count,
            'team_member_data_count': team_member_data_count,
            'database_url': app.config['SQLALCHEMY_DATABASE_URI']
        }), 200
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Database connection failed: {str(e)}',
            'database_url': app.config['SQLALCHEMY_DATABASE_URI']
        }), 500

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, host='0.0.0.0', port=5001) 